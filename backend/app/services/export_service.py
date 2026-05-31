import io
import csv
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models.contact_model import Contact

def generate_contacts_csv(contacts: List[Contact]) -> str:
    """Format contacts into Excel-friendly CSV data."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    writer.writerow([
        "ID", "Full Name", "Email", "Phone", 
        "Company", "Service Required", "Message", 
        "IP Address", "Country", "Status", "Submitted At"
    ])
    
    # Data rows
    for c in contacts:
        writer.writerow([
            c.id,
            c.full_name,
            c.email,
            c.phone,
            c.company or "Not specified",
            c.service,
            c.message,
            c.ip_address or "Unknown",
            c.country or "Unknown",
            c.status,
            c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else ""
        ])
        
    return output.getvalue()

def generate_contacts_excel(contacts: List[Contact]) -> bytes:
    """Format contacts into a styled, professional Excel (.xlsx) file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Submissions"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Headers definition
    headers = [
        "ID", "Full Name", "Email", "Phone", 
        "Company", "Service Required", "Message", 
        "IP Address", "Country", "Status", "Submitted At"
    ]
    ws.append(headers)
    
    # Header Styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Slate gray 800
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align if col_idx in [1, 4, 8, 9, 10, 11] else left_align
        cell.border = thin_border
    
    # Data Population
    data_font = Font(name="Segoe UI", size=10)
    for c in contacts:
        row_data = [
            c.id,
            c.full_name,
            c.email,
            c.phone,
            c.company or "-",
            c.service,
            c.message,
            c.ip_address or "Unknown",
            c.country or "Unknown",
            c.status.upper(),
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""
        ]
        ws.append(row_data)
        
        # Style the populated row
        current_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # Alignments
            if col_idx in [1, 4, 8, 9, 10, 11]:
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col_idx == 7) # Wrap text for message
    
    # Auto-adjust column widths with safety margin
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        
        for cell in col:
            # Skip messages wrap length in width check to avoid ultra-wide column
            if cell.column == 7: 
                max_len = max(max_len, 15)
                continue
            val_str = str(cell.value or '')
            max_len = max(max_len, len(val_str))
            
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)
        
    # Write workbook to bytes stream
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()
